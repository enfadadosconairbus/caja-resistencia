#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
casador.py — Conciliación por NOMBRE + IMPORTE (sin código AIR26)

Para cuando el extracto del banco NO trae el concepto real (AIR26-XXXXX) sino
"trf." + nombre del ordenante (apellidos + nombre). Ver CONCILIACION.md §2/§7.

Casa cada ingreso del banco contra la hoja PEDIDOS por:
  1) NOMBRE normalizado (sin orden, sin acentos, ñ→n, sin "trf.")
  2) IMPORTE exacto al céntimo  ==  TOTAL_EUR del pedido

Reparte en tres cubos:
  CASADO            un único pedido casa por nombre+importe (alta confianza)
  REVISAR_AMBIGUO   varios pedidos posibles, o solo un parecido aproximado
  SIN_MATCH         ningún pedido por nombre (probable pago desde cuenta de un tercero)

NO envía emails ni toca la hoja. Solo produce un Excel para que decidas a mano
(luego: seleccionar fila en PEDIDOS → «✅ Confirmar PAGO del seleccionado»).

Uso:
  python casador.py --banco "Movimientos-conciliacion.xlsx" \
                    --pedidos PEDIDOS.csv \
                    --salida  conciliacion-resultado.xlsx

--pedidos admite .csv (export de la hoja PEDIDOS) o .xlsx.
Reconoce las columnas por su cabecera: ID, NOMBRE, APELLIDOS, EMAIL, TOTAL_EUR, ESTADO.
"""
import argparse
import csv
import re
import sys
import unicodedata
from difflib import SequenceMatcher

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Estados que ya cuentan como pagados (espejo de ESTADOS_PAGADOS en Code.gs).
ESTADOS_PAGADOS = {"PAGO_CONCILIADO", "ENVIADO_PROVEEDOR", "LISTO_RECOGIDA",
                   "ENTREGADO", "YA_PAGADO"}

# Umbral de parecido para el candidato aproximado (0..1) cuando no hay match exacto.
UMBRAL_FUZZY = 0.86

# Conectores que no aportan a la identidad; se ignoran al comparar el juego de tokens.
CONECTORES = {"de", "del", "la", "las", "los", "y", "e", "da", "do"}


def normalizar(texto):
    """minúsculas, sin 'trf.', sin acentos, ñ→n, solo letras/espacios."""
    s = str(texto or "").lower()
    s = re.sub(r"^\s*trf\.?\s*", "", s)          # quita el prefijo del banco
    s = s.replace("transferencia", " ")
    s = s.replace("�", "n")                 # ñ rota (mu�oz) → n
    # descompone y elimina marcas: á→a, ñ→n, ü→u ...
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9 ]", " ", s)            # fuera puntuación
    s = re.sub(r"\s+", " ", s).strip()
    return s


def tokens_clave(texto):
    """conjunto de tokens significativos (sin conectores) para comparar sin orden."""
    return frozenset(t for t in normalizar(texto).split() if t not in CONECTORES)


def compacto(texto):
    """tokens ORDENADOS y pegados, para el parecido aproximado sin depender del orden
    (tolera 'julio vadillo mu oz' vs 'muñoz vadillo julio')."""
    return "".join(sorted(normalizar(texto).split()))


def cent(x):
    return round(float(x) * 100)


def parse_importe(v):
    """espejo de parseImporte(): admite '1.234,56', '10,00', '10.00', número."""
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d,.\-]", "", str(v or ""))
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s or 0)
    except ValueError:
        return 0.0


# ----------------------------- lectura de datos -----------------------------

def leer_banco(path):
    """Excel generado desde el PDF: FECHA OPERACIÓN, FECHA VALOR, CONCEPTO, IMPORTE, SALDO."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    cab = [str(c or "").strip().upper() for c in filas[0]]

    def idx(*claves):
        for k in claves:
            for i, c in enumerate(cab):
                if k in c:
                    return i
        return None

    i_con = idx("CONCEPTO")
    i_imp = idx("IMPORTE")
    i_fop = idx("OPERACI", "FECHA DE LA")
    out = []
    for r in filas[1:]:
        if r is None or i_con is None or r[i_con] in (None, ""):
            continue
        out.append({
            "fecha": r[i_fop] if i_fop is not None else "",
            "concepto": str(r[i_con]),
            "importe": parse_importe(r[i_imp]),
        })
    return out


def leer_pedidos(path):
    """CSV o XLSX con cabecera de HEAD.PEDIDOS. Devuelve lista de dicts."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(path, data_only=True)
        # en el libro completo, usa la hoja PEDIDOS; si no, la activa.
        ws = wb["PEDIDOS"] if "PEDIDOS" in wb.sheetnames else wb.active
        filas = [list(r) for r in ws.iter_rows(values_only=True)]
        cab = [str(c or "").strip().upper() for c in filas[0]]
        cuerpo = filas[1:]
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            # autodetecta ; o ,
            muestra = f.read(2048)
            f.seek(0)
            delim = ";" if muestra.count(";") > muestra.count(",") else ","
            rows = list(csv.reader(f, delimiter=delim))
        cab = [str(c or "").strip().upper() for c in rows[0]]
        cuerpo = rows[1:]

    def col(*claves):
        for k in claves:
            for i, c in enumerate(cab):
                if c == k:
                    return i
        for k in claves:                      # segundo intento: subcadena
            for i, c in enumerate(cab):
                if k in c:
                    return i
        return None

    c_id = col("ID")
    c_nom = col("NOMBRE")
    c_ape = col("APELLIDOS")
    c_email = col("EMAIL")
    c_total = col("TOTAL_EUR", "TOTAL")
    c_estado = col("ESTADO")

    if c_nom is None or c_total is None:
        sys.exit("El export de PEDIDOS no tiene columnas NOMBRE / TOTAL_EUR reconocibles.\n"
                 "Cabecera leída: " + ", ".join(cab))

    peds = []
    for r in cuerpo:
        if not r or (c_nom < len(r) and (r[c_nom] in (None, ""))):
            continue
        nombre = str(r[c_nom]) if c_nom is not None and c_nom < len(r) else ""
        apell = str(r[c_ape]) if c_ape is not None and c_ape < len(r) and r[c_ape] else ""
        full = (nombre + " " + apell).strip()
        peds.append({
            "id": str(r[c_id]) if c_id is not None and c_id < len(r) else "",
            "nombre_completo": full,
            "email": str(r[c_email]) if c_email is not None and c_email < len(r) and r[c_email] else "",
            "total": parse_importe(r[c_total]) if c_total < len(r) else 0.0,
            "estado": (str(r[c_estado]).strip().upper() if c_estado is not None and c_estado < len(r) and r[c_estado] else ""),
            "tokens": None,   # se rellena abajo
            "compacto": None,
        })
    for p in peds:
        p["tokens"] = tokens_clave(p["nombre_completo"])
        p["compacto"] = compacto(p["nombre_completo"])
    return peds


# ------------------------------- casador ------------------------------------

def casar(banco, pedidos):
    # índice por importe (en céntimos) para acotar candidatos
    por_importe = {}
    for p in pedidos:
        por_importe.setdefault(cent(p["total"]), []).append(p)

    resultados = []
    usados = set()   # ids de pedido ya casados con confianza (evita doble asignación)

    for mov in banco:
        btok = tokens_clave(mov["concepto"])
        bcomp = compacto(mov["concepto"])
        cand = por_importe.get(cent(mov["importe"]), [])

        # 1) match exacto: mismo juego de tokens (sin orden) + importe exacto
        exactos = [p for p in cand if p["tokens"] == btok and btok]
        # 2) si no, subconjunto (banco puede traer un apellido menos o de más)
        if not exactos:
            exactos = [p for p in cand
                       if btok and (btok <= p["tokens"] or p["tokens"] <= btok)
                       and min(len(btok), len(p["tokens"])) >= 2]

        libres = [p for p in exactos if p["id"] not in usados]
        pool = libres if libres else exactos

        if len(pool) == 1:
            p = pool[0]
            estado_pagado = p["estado"] in ESTADOS_PAGADOS
            res = "YA_PAGADO" if estado_pagado else "CASADO"
            if not estado_pagado:
                usados.add(p["id"])
            resultados.append(_fila(mov, res, [p]))
            continue

        if len(pool) > 1:
            resultados.append(_fila(mov, "REVISAR_AMBIGUO", pool,
                                    nota=f"{len(pool)} pedidos con ese nombre e importe"))
            continue

        # 3) sin exacto: candidato aproximado con MISMO importe (tolera ñ rota / 'mu oz')
        aprox = []
        for p in cand:
            score = SequenceMatcher(None, bcomp, p["compacto"]).ratio()
            if score >= UMBRAL_FUZZY:
                aprox.append((score, p))
        aprox.sort(key=lambda x: -x[0])
        if aprox:
            mejores = [p for _, p in aprox[:3]]
            top = aprox[0][0]
            resultados.append(_fila(mov, "REVISAR_AMBIGUO", mejores,
                                    nota=f"parecido {top:.0%} (mismo importe, nombre no idéntico)"))
            continue

        # 4) nada por nombre. ¿al menos hay pedidos con ese importe?
        n_mismo_importe = len(cand)
        nota = (f"{n_mismo_importe} pedido(s) con ese importe pero otro nombre "
                "— posible pago desde cuenta de un tercero") if n_mismo_importe else \
               "ningún pedido con ese importe"
        resultados.append(_fila(mov, "SIN_MATCH", [], nota=nota))

    return resultados


def _fila(mov, resultado, candidatos, nota=""):
    c = candidatos[0] if candidatos else None
    return {
        "fecha": mov["fecha"],
        "concepto": mov["concepto"],
        "importe": mov["importe"],
        "resultado": resultado,
        "pedido_id": c["id"] if c else "",
        "pedido_nombre": c["nombre_completo"] if c else "",
        "pedido_email": c["email"] if c else "",
        "pedido_total": c["total"] if c else "",
        "otros_candidatos": " | ".join(f'{p["id"]}·{p["nombre_completo"]}'
                                       for p in candidatos[1:]) if len(candidatos) > 1 else "",
        "nota": nota,
    }


# ------------------------------- salida -------------------------------------

def escribir(resultados, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliación"
    cols = ["FECHA", "CONCEPTO (banco)", "IMPORTE", "RESULTADO", "PEDIDO_ID",
            "PEDIDO_NOMBRE", "PEDIDO_EMAIL", "PEDIDO_TOTAL", "OTROS_CANDIDATOS", "NOTA"]
    ws.append(cols)
    colores = {
        "CASADO": "C6EFCE", "YA_PAGADO": "DDEBF7",
        "REVISAR_AMBIGUO": "FFEB9C", "SIN_MATCH": "FFC7CE",
    }
    hfill = PatternFill("solid", fgColor="1F3A5F")
    for i in range(1, len(cols) + 1):
        cell = ws.cell(1, i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    # orden: primero lo que hay que revisar, luego lo casado
    orden = {"REVISAR_AMBIGUO": 0, "SIN_MATCH": 1, "YA_PAGADO": 2, "CASADO": 3}
    resultados = sorted(resultados, key=lambda r: orden.get(r["resultado"], 9))

    for r in resultados:
        ws.append([r["fecha"], r["concepto"], r["importe"], r["resultado"],
                   r["pedido_id"], r["pedido_nombre"], r["pedido_email"],
                   r["pedido_total"], r["otros_candidatos"], r["nota"]])
        fila = ws.max_row
        color = colores.get(r["resultado"])
        if color:
            ws.cell(fila, 4).fill = PatternFill("solid", fgColor=color)
        ws.cell(fila, 3).number_format = "#,##0.00 €"
        if r["pedido_total"] != "":
            ws.cell(fila, 8).number_format = "#,##0.00 €"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    for i, w in enumerate([12, 42, 12, 18, 14, 30, 30, 14, 40, 46], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # hoja resumen
    rs = wb.create_sheet("Resumen")
    from collections import Counter
    cuenta = Counter(r["resultado"] for r in resultados)
    suma = Counter()
    for r in resultados:
        suma[r["resultado"]] += r["importe"]
    rs.append(["RESULTADO", "Nº MOVIMIENTOS", "IMPORTE (€)"])
    for k in ["CASADO", "YA_PAGADO", "REVISAR_AMBIGUO", "SIN_MATCH"]:
        rs.append([k, cuenta.get(k, 0), round(suma.get(k, 0.0), 2)])
    rs.append(["TOTAL", sum(cuenta.values()), round(sum(suma.values()), 2)])
    for i in range(1, 4):
        rs.cell(1, i).font = Font(bold=True)
    rs.column_dimensions["A"].width = 20
    rs.column_dimensions["B"].width = 16
    rs.column_dimensions["C"].width = 14

    wb.save(path)
    return cuenta, suma


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")   # consola Windows (cp1252) → UTF-8
    except Exception:
        pass
    ap = argparse.ArgumentParser(description="Conciliación banco↔pedidos por nombre+importe")
    ap.add_argument("--banco", required=True, help="Excel de movimientos (generado del PDF)")
    ap.add_argument("--pedidos", required=True, help="Export de la hoja PEDIDOS (.csv o .xlsx)")
    ap.add_argument("--salida", default="conciliacion-resultado.xlsx")
    a = ap.parse_args()

    banco = leer_banco(a.banco)
    pedidos = leer_pedidos(a.pedidos)
    resultados = casar(banco, pedidos)
    cuenta, suma = escribir(resultados, a.salida)

    print(f"Movimientos leídos : {len(banco)}")
    print(f"Pedidos leídos     : {len(pedidos)}")
    print("-" * 40)
    for k in ["CASADO", "YA_PAGADO", "REVISAR_AMBIGUO", "SIN_MATCH"]:
        print(f"  {k:<16} {cuenta.get(k,0):>4}   {round(suma.get(k,0.0),2):>10.2f} €")
    print("-" * 40)
    print(f"Salida → {a.salida}")


if __name__ == "__main__":
    main()
