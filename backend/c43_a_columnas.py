#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
c43_a_columnas.py — Decodifica un extracto bancario Norma 43 (Cuaderno 43 / AEB43)
a un Excel con una columna por dato.

A diferencia del PDF/Excel "de pantalla" del banco (que recorta el concepto al
nombre del ordenante), el C43 trae el CONCEPTO REAL en los registros de tipo 23:
el nombre del ordenante Y el código de pedido (AIR26-XXXXX). Con eso la conciliación
puede hacerse por CÓDIGO + importe (exacta), sin depender del cruce por nombre.

Formato (registros de 80 caracteres):
  11  cabecera de cuenta   (entidad, oficina, cuenta, fechas, saldo inicial)
  22  movimiento           (fechas, debe/haber, importe)
  23  conceptos complementarios de un 22: [4:42] campo 1, [42:80] campo 2
  33  fin de cuenta        (nº apuntes, totales, saldo final)
  88  fin de fichero

El fichero de entrada puede ser el .txt del C43, o un .docx que lo envuelva (a
veces llega así); ambos se leen igual.

Uso:
  python c43_a_columnas.py --entrada "C43.TXT" --salida C43-en-columnas.xlsx

Salida (hoja "Movimientos C43"):
  ORDEN · FECHA OPERACIÓN · FECHA VALOR · NOMBRE ORDENANTE ·
  CONCEPTO (texto banco) · CÓDIGO PEDIDO · IMPORTE · SALDO
Los movimientos sin código AIR26 detectable se marcan en rojo (donativos con
mensaje, transferencias desde Revolut/extranjero, códigos mal escritos…).
"""
import argparse
import datetime
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def leer_lineas(path):
    """Devuelve las líneas del C43, tanto si es .txt como si viene en un .docx."""
    if path.lower().endswith((".docx",)):
        import docx
        d = docx.Document(path)
        return [p.text for p in d.paragraphs if p.text.strip()]
    # texto plano: prueba varias codificaciones habituales de los bancos
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            with open(path, encoding=enc) as f:
                return [ln.rstrip("\n\r") for ln in f if ln.strip()]
        except UnicodeDecodeError:
            continue
    with open(path, encoding="utf-8", errors="replace") as f:
        return [ln.rstrip("\n\r") for ln in f if ln.strip()]


def fecha(aammdd):
    if not re.match(r"\d{6}$", aammdd):
        return None
    return datetime.date(2000 + int(aammdd[:2]), int(aammdd[2:4]), int(aammdd[4:6]))


def norm_cod(texto):
    """Mismo criterio que detectarId() del Code.gs: AIR26 + número real → AIR26-XXXXX."""
    flat = re.sub(r"[^A-Z0-9]", "", str(texto).upper())
    m = re.search(r"AIR26" + r"0*(\d{1,6})", flat)
    return "AIR26-" + str(int(m.group(1))).zfill(5) if m else ""


def parse(lineas):
    movs, cur = [], None
    for x in lineas:
        t = x[:2]
        if t == "22":
            imp = int(x[28:42]); dh = x[27]
            cur = {"fop": fecha(x[10:16]), "fval": fecha(x[16:22]),
                   "imp": imp / 100.0 * (1 if dh == "2" else -1),
                   "nombre": "", "texto": ""}
            movs.append(cur)
        elif t == "23" and cur is not None:
            cur["nombre"] = (cur["nombre"] + " " + x[4:42].strip()).strip()
            cur["texto"] = (cur["texto"] + " " + x[42:80].strip()).strip()
    return movs


def escribir(movs, out):
    wb = Workbook(); ws = wb.active; ws.title = "Movimientos C43"
    cab = ["ORDEN", "FECHA OPERACIÓN", "FECHA VALOR", "NOMBRE ORDENANTE",
           "CONCEPTO (texto banco)", "CÓDIGO PEDIDO", "IMPORTE", "SALDO"]
    ws.append(cab)
    hfill = PatternFill("solid", fgColor="1F3A5F")
    for i in range(1, len(cab) + 1):
        c = ws.cell(1, i); c.font = Font(bold=True, color="FFFFFF"); c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    saldo, sin_cod = 0.0, 0
    for i, m in enumerate(movs, 1):
        saldo += m["imp"]
        cod = norm_cod(m["texto"])
        if not cod:
            sin_cod += 1
        ws.append([i, m["fop"], m["fval"], m["nombre"], m["texto"], cod, m["imp"], round(saldo, 2)])
        r = ws.max_row
        ws.cell(r, 2).number_format = "DD/MM/YYYY"
        ws.cell(r, 3).number_format = "DD/MM/YYYY"
        ws.cell(r, 7).number_format = "#,##0.00 €"
        ws.cell(r, 8).number_format = "#,##0.00 €"
        if not cod:
            ws.cell(r, 6).fill = PatternFill("solid", fgColor="FFC7CE")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    for i, w in enumerate([8, 18, 14, 34, 40, 16, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(out)
    return sin_cod


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    ap = argparse.ArgumentParser(description="Decodifica un C43 (Norma 43) a Excel por columnas")
    ap.add_argument("--entrada", required=True, help="Fichero C43 (.txt o .docx que lo envuelva)")
    ap.add_argument("--salida", default="C43-en-columnas.xlsx")
    a = ap.parse_args()

    movs = parse(leer_lineas(a.entrada))
    sin_cod = escribir(movs, a.salida)
    print(f"Movimientos     : {len(movs)}")
    print(f"Suma importes   : {round(sum(m['imp'] for m in movs), 2)} €")
    print(f"Con código AIR26: {len(movs) - sin_cod}  |  sin código: {sin_cod}")
    print(f"Salida → {a.salida}")


if __name__ == "__main__":
    main()
